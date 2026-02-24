#!/usr/bin/env python3
"""Import MCTV Host List spreadsheet into Supabase.

Imports:
  - Host List sheet → clients table (client_type='host')
  - Advertiser List sheet → clients table (client_type='advertiser')
  - Hot List sheet → leads table

Source: MCTV Host List (1).xlsx
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env", override=True)

import openpyxl
from services.supabase_client import insert_row, query_table

XLSX_PATH = Path(r"C:\Users\msaac\Downloads\MCTV Host List (1).xlsx")

# Category normalization for industry field
CATEGORY_MAP = {
    "Bar/Restaurant": "Food & Beverage",
    "Retail": "Retail",
    "Barbershop/Salon": "Beauty & Wellness",
    "Hair Salon": "Beauty & Wellness",
    "Medical": "Healthcare",
    "Other": "Other",
    "Family Rec & Entertainment": "Entertainment",
    "Education": "Education",
    "Liquor/Wine/Beer Store": "Retail",
    "Health & Fitness": "Health & Fitness",
    "Gas/ Grocery": "Retail",
    "Non Profit, Community, Government": "Nonprofit",
    "Professional Services": "Professional Services",
    "Travel & Tourism": "Hospitality",
    "Auto Shop/Auto Dealer/Oil Change": "Automotive",
    "Coffee/Donut/Bagel Shop": "Food & Beverage",
}


def import_hosts(wb) -> int:
    """Import Host List sheet into clients table as hosts."""
    ws = wb["Host List"]
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name:
            continue

        category = str(row[1] or "").strip()
        address = str(row[2] or "").strip()
        city = str(row[3] or "").strip()
        state = str(row[4] or "").strip()
        postal = str(row[5] or "").replace(".0", "").strip()
        licenses = int(float(row[6])) if row[6] else 1
        first = str(row[7] or "").strip()
        last = str(row[8] or "").strip()
        contact_type = str(row[9] or "").strip()
        email = str(row[10] or "").strip()
        phone = str(row[11] or "").strip()

        contact_name = f"{first} {last}".strip() or "On File"

        industry = CATEGORY_MAP.get(category, category)

        # Build notes with extra info
        notes_parts = []
        if address:
            notes_parts.append(f"Address: {address}, {city}, {state} {postal}")
        if contact_type:
            notes_parts.append(f"Contact type: {contact_type}")
        if category:
            notes_parts.append(f"Venue category: {category}")
        if licenses > 1:
            notes_parts.append(f"Screen count: {licenses}")

        data = {
            "business_name": name,
            "contact_name": contact_name,
            "contact_email": email or f"",
            "contact_phone": phone,
            "industry": industry,
            "city": city,
            "client_type": "host",
            "status": "active",
            "notes": " | ".join(notes_parts) if notes_parts else None,
        }

        result = insert_row("clients", data)
        if result:
            count += 1
            print(f"  [+] {name} ({city}, {licenses} screen{'s' if licenses > 1 else ''}) - {industry}")
        else:
            print(f"  [!] Failed: {name}")

    return count


def import_advertisers(wb) -> int:
    """Import Advertiser List sheet into clients table as advertisers."""
    ws = wb["Advertiser List"]
    count = 0
    seen = set()  # Avoid duplicates (Pregnancy Center of Oxford listed twice)

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)

        category = str(row[1] or "").strip()
        address = str(row[2] or "").strip()
        city = str(row[3] or "").strip()
        state = str(row[4] or "").strip()
        first = str(row[5] or "").strip()
        last = str(row[6] or "").strip()
        title = str(row[7] or "").strip()
        email = str(row[8] or "").strip()
        phone = str(row[9] or "").strip()

        contact_name = f"{first} {last}".strip() or "On File"

        notes_parts = []
        if address:
            notes_parts.append(f"Address: {address}, {city}, {state}")
        if category:
            notes_parts.append(f"Business type: {category}")
        if title:
            notes_parts.append(f"Title: {title}")

        data = {
            "business_name": name,
            "contact_name": contact_name,
            "contact_email": email or f"",
            "contact_phone": phone,
            "industry": category or "Other",
            "city": city,
            "client_type": "advertiser",
            "status": "active",
            "notes": " | ".join(notes_parts) if notes_parts else None,
        }

        result = insert_row("clients", data)
        if result:
            count += 1
            print(f"  [+] {name} ({city}) - {category}")
        else:
            print(f"  [!] Failed: {name}")

    return count


def import_hot_list(wb) -> int:
    """Import Hot List sheet into leads table."""
    ws = wb["Hot List"]
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name:
            continue

        category = str(row[1] or "").strip()
        address = str(row[2] or "").strip()
        city = str(row[3] or "").strip()
        state = str(row[4] or "").strip()
        first = str(row[5] or "").strip()
        last = str(row[6] or "").strip()
        title = str(row[7] or "").strip()
        email = str(row[8] or "").strip()
        phone = str(row[9] or "").strip()

        contact_name = f"{first} {last}".strip() or ""

        data = {
            "business_name": name,
            "contact_name": contact_name,
            "contact_email": email,
            "contact_phone": phone,
            "industry": category or "",
            "city": city or "",
            "interest_level": "warm",
            "status": "new",
            "additional_notes": f"Imported from MCTV Hot List. {f'Title: {title}' if title else ''}".strip(),
        }

        result = insert_row("leads", data)
        if result:
            count += 1
            print(f"  [+] {name} {f'({city})' if city else ''} {f'- {contact_name}' if contact_name else ''}")
        else:
            print(f"  [!] Failed: {name}")

    return count


def main():
    print("=" * 60)
    print("  MCTV HOST LIST IMPORT")
    print("=" * 60)

    if not XLSX_PATH.exists():
        print(f"[!] File not found: {XLSX_PATH}")
        return 1

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Import hosts
    print()
    print("--- Importing Host Venues (→ clients table) ---")
    host_count = import_hosts(wb)

    # Import advertisers
    print()
    print("--- Importing Advertisers (→ clients table) ---")
    adv_count = import_advertisers(wb)

    # Import hot list
    print()
    print("--- Importing Hot List (→ leads table) ---")
    hot_count = import_hot_list(wb)

    # Summary
    print()
    print("=" * 60)
    print(f"  IMPORT COMPLETE")
    print(f"    Hosts:       {host_count} venues imported")
    print(f"    Advertisers: {adv_count} clients imported")
    print(f"    Hot List:    {hot_count} leads imported")
    print(f"    Total:       {host_count + adv_count + hot_count} records")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
