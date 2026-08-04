# Copyright (c) 2026 MCTV Digital, Inc. All rights reserved.
# Proprietary and confidential. Unauthorized copying, distribution,
# or modification of this file is strictly prohibited.
"""Field audit service — the screen roster a technician can actually work from.

Everything needed to send someone into the field lives in five places and
nothing joins them. This does:

    screen_loops ............ the spine. One row per license (Raspberry Pi) per
                              sweep date, from the n-compass whitelist sweep.
                              This is where license numbers come from.
    venue_rate_inputs ....... screen count and venue type, and it is the only
                              table that knows about `special`-market venues
                              (the airport) that the market sweeps miss.
    clients ................. host contact name / phone / email, plus the
                              street address buried in the notes string.
    network_dashboard.json .. NTV360's address and category per venue.
    venue_geocodes.json ..... lat/lon, and a USPS-normalised address that
                              corrects a couple of the dashboard ones.

Venue names are spelled differently in all five, so every join goes through
``screen_inventory_service.normalize``.

Public API:
    short_code(market, license_id) .... stable label code, e.g. "TUP-6ABE3B"
    audit_roster(market, sweep) ....... one dict per screen, fully joined
    route_order(rows) ................. nearest-neighbour + 2-opt, adds
                                        stop_no / cluster / same_building
    roster_gaps(rows) ................. what the packet is missing and why
    build_workbook(rows, gaps, market)  the field sheet, as .xlsx bytes

Screens that n-compass knows about but the sweep never captured live in
``data/audit/extra_licenses.json`` and are merged in with ``source='overlay'``.
"""

from __future__ import annotations

import json
import math
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from services.screen_inventory_service import (
    is_diagnostic,
    market_label,
    mmss,
    normalize,
)

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
OVERLAY_PATH = DATA_DIR / "audit" / "extra_licenses.json"

# 15:00 loop target (4 plays/hr) — the same target the sweep is measured against.
TARGET_SECONDS = 900

# Short-code prefix per market. Anything unlisted falls back to the first three
# letters of the market name.
SHORT_CODE_PREFIX = {
    "tupelo": "TUP",
    "oxford": "OXF",
    "starkville": "STK",
    "columbus": "CLB",
    "west_point": "WPT",
    "special": "SPC",
}

# Consecutive stops closer together than this join the same cluster — one leg
# of the day rather than two separate drives.
CLUSTER_GAP_MI = 0.75

# ...but a cluster stops growing once it spans more than this from its own
# centre. Without it a dense strip chains end to end and "one leg" quietly
# becomes half the town.
CLUSTER_SPAN_MI = 0.6

# Street-type suffixes. Everything after one of these in an address is a suite
# or unit designator, which is exactly what has to come off before two tenants
# in one building will match each other.
STREET_TYPES = {
    "st", "street", "ave", "avenue", "blvd", "boulevard", "rd", "road", "dr",
    "drive", "ln", "lane", "way", "pkwy", "parkway", "hwy", "highway", "ct",
    "court", "pl", "place", "cir", "circle", "ter", "terrace", "trl", "trail",
    "sq", "square", "loop", "run", "pike", "row",
}

SUITE_MARKERS = {"ste", "suite", "unit", "apt", "apartment", "#", "no", "bldg", "building"}

EARTH_RADIUS_MI = 3958.8


# ── Columns the field sheet captures ─────────────────────────────────────────
#
# (header, width, dropdown options or None). Order is the order they appear in
# the workbook, and the DOCX brief's legend renders from this same list so the
# two can never drift apart.

CAPTURE_COLUMNS: list[tuple[str, int, list[str] | None]] = [
    ("Screen on?",          12, ["Yes", "No", "No screen found"]),
    ("Content playing?",    16, ["Yes", "Frozen", "Black", "Error / no signal"]),
    ("Display brand",       16, None),
    ("Display model",       18, None),
    ("Size (in)",           10, None),
    ("Orientation",         13, ["Landscape", "Portrait"]),
    ("Mount type",          16, ["Wall", "Ceiling", "Stand", "Shelf / counter", "Other"]),
    ("Pi model",            16, ["Pi 3", "Pi 4", "Pi 5", "Other / unmarked"]),
    ("Pi serial",           20, None),
    ("Power source",        18, ["Wall outlet", "Power strip", "TV USB port", "PoE", "Other"]),
    ("Network",             16, ["Wi-Fi", "Ethernet", "Unknown"]),
    ("Label applied?",      14, ["Yes", "No - Pi not reachable", "No - no label for this screen"]),
    ("Condition",           14, ["Good", "Fair", "Poor", "Needs replacement"]),
    ("Issues found",        34, None),
    ("Photo taken?",        13, ["Yes", "No"]),
    ("Audited by",          16, None),
    ("Date",                12, None),
    ("Notes",               40, None),
]

# What each capture column is for. Printed in the brief's legend and the
# workbook's Legend sheet.
CAPTURE_LEGEND: dict[str, str] = {
    "Screen on?": "Is the display powered and showing something? 'No screen found' if the venue has no MCTV screen at all.",
    "Content playing?": "Is the MCTV loop actually advancing? Watch for at least two spot changes before answering.",
    "Display brand": "Manufacturer on the bezel or the back label (Samsung, LG, TCL, Vizio...).",
    "Display model": "Model number from the back label. Photograph it if it is hard to read.",
    "Size (in)": "Diagonal screen size in inches. The model number usually starts with it.",
    "Orientation": "How the display is hung. Our creative is built landscape; portrait screens letterbox.",
    "Mount type": "How it is fixed in place. Tells us what it costs to service or replace.",
    "Pi model": "Raspberry Pi generation, printed on the board. 'Other / unmarked' is a fine answer.",
    "Pi serial": "Serial from the Pi's own label, not the license number. Blank if the unit cannot be reached.",
    "Power source": "What the Pi is plugged into. TV USB ports cut power with the TV and cause 'dark screen' tickets.",
    "Network": "How the Pi connects. Note the SSID in Notes if it is on venue guest Wi-Fi.",
    "Label applied?": "Confirm the MCTV license label was affixed to this unit.",
    "Condition": "Overall physical state of display and Pi together.",
    "Issues found": "Anything wrong: dark screen, wrong content, damaged mount, cables exposed, blocked sightline.",
    "Photo taken?": "One wide shot of the screen in place, one close shot of the Pi with its new label.",
    "Audited by": "Technician name.",
    "Date": "Date of the visit (YYYY-MM-DD).",
    "Notes": "Anything the columns above do not cover — including who you spoke to.",
}

# Workbook column -> screen_assets column. Anything not listed here is context
# for the technician and is not stored.
CAPTURE_DB_FIELDS: dict[str, str] = {
    "Screen on?": "screen_on",
    "Content playing?": "content_playing",
    "Display brand": "display_brand",
    "Display model": "display_model",
    "Size (in)": "display_size_in",
    "Orientation": "orientation",
    "Mount type": "mount_type",
    "Pi model": "pi_model",
    "Pi serial": "pi_serial",
    "Power source": "power_source",
    "Network": "network_type",
    "Label applied?": "label_applied",
    "Condition": "condition",
    "Issues found": "issues",
    "Photo taken?": "photo_taken",
    "Audited by": "audited_by",
    "Date": "audited_at",
    "Notes": "notes",
}

# Pre-filled columns, in workbook order: (header, roster key, width).
PREFILL_COLUMNS: list[tuple[str, str, int]] = [
    ("Stop",           "stop_no",               6),
    ("Cluster",        "cluster",              22),
    ("Venue",          "venue_name",           34),
    ("Screen",         "screen_label",         16),
    ("Address",        "address",              38),
    ("Contact",        "contact_name",         20),
    ("Phone",          "contact_phone",        16),
    ("Short code",     "short_code",           13),
    ("License number", "license_id",           38),
    ("Expected loop",  "expected_loop_mmss",   14),
    ("Expected items", "expected_items",       14),
]


# ── Short codes ──────────────────────────────────────────────────────────────

def short_code(market: str | None, license_id: str | None) -> str:
    """Stable per-screen code for labels and field sheets, e.g. ``TUP-6ABE3B``.

    Derived from the license UUID rather than from position in a list, so it
    stays correct when venues are added or removed, and a technician holding a
    label can match it to the license number by eye. Returns "" when the
    license is not known yet — those screens print with a blank code area.
    """
    if not license_id:
        return ""
    key = (market or "").strip().lower()
    prefix = SHORT_CODE_PREFIX.get(key) or (key[:3].upper() or "MCT")
    hexed = "".join(ch for ch in str(license_id) if ch in "0123456789abcdefABCDEF")
    return f"{prefix}-{hexed[:6].upper()}"


# ── Source loaders ───────────────────────────────────────────────────────────

def _load_json(path: Path) -> Any:
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return None


def _dashboard_by_venue() -> dict[str, dict]:
    """NTV360 network dashboard venues, keyed by normalised host name."""
    blob = _load_json(DATA_DIR / "network_dashboard.json") or {}
    out: dict[str, dict] = {}
    for row in (blob.get("venues") or {}).values():
        key = normalize(row.get("host_name"))
        if key:
            out[key] = row
    return out


def _geocodes_by_venue() -> dict[str, dict]:
    blob = _load_json(DATA_DIR / "venue_geocodes.json") or {}
    return {normalize(name): row for name, row in blob.items() if normalize(name)}


def _parse_host_notes(notes: str | None) -> dict[str, str]:
    """Pull the pipe-delimited fields out of a ``clients.notes`` string.

    Host rows imported from the host list carry their address and role in
    notes rather than in columns:

        "Address: 314 E Main St, Tupelo, MS 38804 | Contact type: Manager |
         Venue category: Travel & Tourism | Screen count: 2"
    """
    out: dict[str, str] = {}
    for part in (notes or "").split("|"):
        label, _, value = part.partition(":")
        label, value = label.strip().lower(), value.strip()
        if not value:
            continue
        if label == "address":
            out["address"] = value
        elif label == "contact type":
            out["contact_role"] = value
        elif label == "venue category":
            out["category"] = value
        elif label == "screen count":
            out["screen_count"] = value
    return out


def _hosts_by_venue() -> dict[str, dict]:
    """Host rows from ``clients``, keyed by normalised business name."""
    try:
        from services.supabase_client import query_table
        rows = query_table("clients", filters={"client_type": "host"})
    except Exception:
        rows = []
    out: dict[str, dict] = {}
    for row in rows or []:
        key = normalize(row.get("business_name"))
        if key:
            out[key] = row
    return out


def _rate_inputs_by_venue(market: str) -> dict[str, dict]:
    """Rate-model inputs for a market, plus any ``special``-market venue in the
    same city. The airport lives in ``special`` but is physically in Tupelo, so
    a market-only filter would silently drop six screens."""
    try:
        from services.supabase_client import query_table
        rows = query_table("venue_rate_inputs")
    except Exception:
        rows = []
    city = market_label(market).lower()
    out: dict[str, dict] = {}
    for row in rows or []:
        in_market = (row.get("market") or "").strip().lower() == market
        in_city = (row.get("city") or "").strip().lower() == city
        if in_market or in_city:
            key = normalize(row.get("venue_name"))
            if key:
                out[key] = row
    return out


def _overlay_rows(market: str) -> list[dict]:
    """Screens n-compass knows about that the sweep never captured.

    Matched on the roster's market, or on the overlay row's ``city`` so a
    ``special``-market venue still lands in its home market's packet.
    """
    rows = _load_json(OVERLAY_PATH) or []
    city = market_label(market).lower()
    return [
        r for r in rows
        if (r.get("market") or "").strip().lower() == market
        or (r.get("city") or "").strip().lower() == city
    ]


# ── Roster ───────────────────────────────────────────────────────────────────

def audit_roster(market: str, sweep: str | None = None) -> list[dict[str, Any]]:
    """Every screen a technician should visit in ``market``, fully joined.

    One row per license. Rows sourced from the whitelist sweep carry a real
    ``license_id``; overlay rows carry ``None`` until the number is pulled from
    n-compass, and render as a blank for the technician to fill in on site.
    """
    from services.loop_inventory_service import latest_sweep_date, screen_loops

    market = (market or "").strip().lower()
    sweep = sweep or latest_sweep_date()

    dashboard = _dashboard_by_venue()
    geocodes = _geocodes_by_venue()
    hosts = _hosts_by_venue()
    rate_inputs = _rate_inputs_by_venue(market)

    sweep_rows = [
        r for r in screen_loops(sweep)
        if (r.get("market") or "").strip().lower() == market
        and not is_diagnostic(r.get("venue_name"))
    ]

    # How many licenses the sweep saw at each venue, so a venue with two Pis
    # says so on every one of its rows.
    sweep_counts: dict[str, int] = {}
    for r in sweep_rows:
        sweep_counts[normalize(r.get("venue_name"))] = (
            sweep_counts.get(normalize(r.get("venue_name")), 0) + 1
        )

    rows: list[dict[str, Any]] = []

    for r in sweep_rows:
        rows.append(_build_row(
            market=market,
            venue_name=r.get("venue_name") or "",
            license_id=r.get("license_id"),
            source="sweep",
            screen_label="",
            expected_loop_seconds=r.get("loop_seconds"),
            expected_items=r.get("item_count"),
            over_target=bool(r.get("over_target")),
            sweep_screens=sweep_counts.get(normalize(r.get("venue_name")), 1),
            dashboard=dashboard, geocodes=geocodes,
            hosts=hosts, rate_inputs=rate_inputs,
            overlay_note="",
        ))

    for o in _overlay_rows(market):
        rows.append(_build_row(
            market=market,
            venue_name=o.get("venue_name") or "",
            license_id=o.get("license_id"),
            source="overlay",
            screen_label=o.get("screen_label") or "",
            expected_loop_seconds=o.get("expected_loop_seconds"),
            expected_items=o.get("expected_items"),
            over_target=False,
            sweep_screens=0,
            dashboard=dashboard, geocodes=geocodes,
            hosts=hosts, rate_inputs=rate_inputs,
            overlay_note=o.get("note") or "",
            address_override=o.get("address") or "",
        ))

    rows.sort(key=lambda r: (r["venue_name"].lower(), r["short_code"]))
    return rows


def _build_row(
    *, market: str, venue_name: str, license_id: str | None, source: str,
    screen_label: str, expected_loop_seconds, expected_items, over_target: bool,
    sweep_screens: int, dashboard: dict, geocodes: dict, hosts: dict,
    rate_inputs: dict, overlay_note: str, address_override: str = "",
) -> dict[str, Any]:
    """Join one screen against every reference source."""
    key = normalize(venue_name)
    dash = dashboard.get(key) or {}
    geo = geocodes.get(key) or {}
    host = hosts.get(key) or {}
    rate = rate_inputs.get(key) or {}
    host_notes = _parse_host_notes(host.get("notes"))

    # The host list is hand-maintained and freshest; the dashboard is next; the
    # geocoder's normalised address is the last resort but the most correct
    # when the other two disagree on a street name.
    address = (
        address_override
        or host_notes.get("address")
        or dash.get("address")
        or geo.get("matched_address")
        or ""
    )

    return {
        "market": market,
        "venue_name": venue_name,
        "license_id": license_id,
        "short_code": short_code(market, license_id),
        "screen_label": screen_label,
        "source": source,
        "address": address,
        "geocoded_address": geo.get("matched_address") or "",
        "lat": geo.get("lat"),
        "lon": geo.get("lon"),
        "contact_name": (host.get("contact_name") or "").strip(),
        "contact_phone": (host.get("contact_phone") or "").strip(),
        "contact_email": (host.get("contact_email") or "").strip(),
        "contact_role": host_notes.get("contact_role", ""),
        "category": dash.get("category") or host_notes.get("category") or "",
        "venue_type": rate.get("type_code") or "",
        "expected_loop_seconds": expected_loop_seconds,
        "expected_loop_mmss": mmss(expected_loop_seconds) if expected_loop_seconds else "",
        "expected_items": expected_items,
        "over_target": over_target,
        "sweep_screens": sweep_screens,
        "rate_screens": rate.get("screens"),
        "dashboard_licenses": dash.get("license_count"),
        "note": overlay_note,
        # Filled in by route_order().
        "stop_no": None,
        "cluster": "",
        "same_building": False,
    }


# ── Routing ──────────────────────────────────────────────────────────────────

def _haversine_mi(a: tuple[float, float], b: tuple[float, float]) -> float:
    lat1, lon1 = math.radians(a[0]), math.radians(a[1])
    lat2, lon2 = math.radians(b[0]), math.radians(b[1])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * EARTH_RADIUS_MI * math.asin(math.sqrt(h))


def _nearest_neighbour(points: list[tuple[float, float]], start: int) -> list[int]:
    order, unvisited = [start], set(range(len(points))) - {start}
    while unvisited:
        last = points[order[-1]]
        nxt = min(unvisited, key=lambda i: _haversine_mi(last, points[i]))
        order.append(nxt)
        unvisited.discard(nxt)
    return order


def _two_opt(points: list[tuple[float, float]], order: list[int]) -> list[int]:
    """Untangle a nearest-neighbour tour.

    Nearest-neighbour reliably strands one or two stops on the wrong side of
    town; 2-opt reverses the segments that cross. Open path, so the last leg
    does not wrap back to the start.
    """
    if len(order) < 4:
        return order
    improved = True
    while improved:
        improved = False
        for i in range(len(order) - 2):
            for j in range(i + 2, len(order) - 1):
                a, b = points[order[i]], points[order[i + 1]]
                c, d = points[order[j]], points[order[j + 1]]
                if (_haversine_mi(a, c) + _haversine_mi(b, d)
                        < _haversine_mi(a, b) + _haversine_mi(c, d) - 1e-9):
                    order[i + 1:j + 1] = reversed(order[i + 1:j + 1])
                    improved = True
    return order


def _street_of(address: str) -> str:
    """Street name from a street address, minus house number and suite.

    Source casing is kept as-is — title-casing turns "McCullough" into
    "Mccullough", and the technician is reading these off a street sign.
    """
    head = (address or "").split(",")[0].strip()
    parts = head.split()
    if parts and any(ch.isdigit() for ch in parts[0]):
        parts = parts[1:]
    out: list[str] = []
    for word in parts:
        bare = word.lower().rstrip(".")
        if bare in SUITE_MARKERS:
            break
        out.append(word)
        if bare in STREET_TYPES:
            break
    return " ".join(out)


def _building_key(address: str) -> str:
    """Bucket key for "is this the same building?".

    Truncates at the street-type suffix, so ``499 S Gloster St B-2`` and
    ``499 S Gloster St Suite F-9`` both reduce to ``499 s gloster st``. Without
    that, four tenants of one building read as four separate stops.
    """
    head = normalize((address or "").split(",")[0])
    if not head:
        return ""
    out: list[str] = []
    for word in head.split():
        if word in SUITE_MARKERS:
            break
        out.append(word)
        if word in STREET_TYPES:
            break
    return " ".join(out)


def route_order(rows: list[dict[str, Any]], start: tuple[float, float] | None = None
                ) -> list[dict[str, Any]]:
    """Order the roster into a drivable day and group it into clusters.

    Geocoded stops are sequenced nearest-neighbour then improved with 2-opt,
    starting from the stop farthest from the market centroid — starting at an
    edge of the territory beats starting in the middle. Screens at the same
    venue stay together, and stops with no coordinates are appended at the end
    under an "Address needed" cluster rather than being dropped.

    Mutates and returns ``rows``.
    """
    by_venue: dict[str, list[dict]] = {}
    for r in rows:
        by_venue.setdefault(normalize(r["venue_name"]), []).append(r)

    located, unlocated = [], []
    for key, group in by_venue.items():
        coords = next(((g["lat"], g["lon"]) for g in group
                       if g.get("lat") is not None and g.get("lon") is not None), None)
        (located if coords else unlocated).append((key, group, coords))

    ordered: list[tuple[str, list[dict], tuple[float, float] | None]] = []
    if located:
        points = [c for _, _, c in located]
        if start is not None:
            begin = min(range(len(points)), key=lambda i: _haversine_mi(start, points[i]))
        else:
            centroid = (sum(p[0] for p in points) / len(points),
                        sum(p[1] for p in points) / len(points))
            begin = max(range(len(points)), key=lambda i: _haversine_mi(centroid, points[i]))
        seq = _two_opt(points, _nearest_neighbour(points, begin))
        ordered = [located[i] for i in seq]

    # Cluster: break the route wherever the next stop is a real drive away, or
    # where the cluster has already spread too wide to call one leg.
    clusters: list[list[tuple[str, list[dict], tuple[float, float]]]] = []
    for entry in ordered:
        if clusters:
            current = clusters[-1]
            centre = (sum(c[2][0] for c in current) / len(current),
                      sum(c[2][1] for c in current) / len(current))
            near = _haversine_mi(current[-1][2], entry[2]) <= CLUSTER_GAP_MI
            tight = _haversine_mi(centre, entry[2]) <= CLUSTER_SPAN_MI
            if near and tight:
                current.append(entry)
                continue
        clusters.append([entry])

    stop_no = 0
    for idx, cluster in enumerate(clusters, start=1):
        # Name the cluster after the street it mostly sits on. When no single
        # street carries most of it, name the top two — "S Gloster St" over a
        # cluster that is half downtown tells the technician the wrong thing.
        streets = [s for s in (_street_of(g[0]["address"]) for _, g, _ in cluster) if s]
        ranked = sorted(set(streets), key=lambda s: (-streets.count(s), s))
        name = ""
        if ranked:
            top = ranked[0]
            if streets.count(top) * 10 >= len(cluster) * 7:
                name = top
            elif len(ranked) > 1:
                name = f"{top} & {ranked[1]}"
            else:
                name = top
        label = f"{idx}. {name}" if name else f"{idx}. Stop group {idx}"
        for _, group, _ in cluster:
            stop_no += 1
            for screen in group:
                screen["stop_no"] = stop_no
                screen["cluster"] = label

    # Stops with no coordinates go last. A venue with an address we simply
    # never geocoded is a different problem from one with no address at all,
    # and the technician can still drive the first kind.
    for _, group, _ in sorted(unlocated, key=lambda e: not e[1][0]["address"]):
        stop_no += 1
        label = ("Route by hand" if group[0]["address"] else "Address needed")
        for screen in group:
            screen["stop_no"] = stop_no
            screen["cluster"] = label

    _flag_same_building(rows)
    rows.sort(key=lambda r: (r["stop_no"] or 0, r["venue_name"].lower(), r["short_code"]))
    return rows


def _flag_same_building(rows: list[dict[str, Any]]) -> None:
    """Mark screens that share a street address with a different venue.

    499 S Gloster St in Tupelo is four licenses across four businesses in one
    building. A technician who does not know that drives it four times.
    """
    buckets: dict[str, set[str]] = {}
    for r in rows:
        key = _building_key(r.get("address"))
        if key:
            buckets.setdefault(key, set()).add(normalize(r["venue_name"]))
    for r in rows:
        r["same_building"] = len(buckets.get(_building_key(r.get("address")), set())) > 1


# ── Gaps ─────────────────────────────────────────────────────────────────────

def roster_gaps(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Everything about this roster that a technician or we should know is
    incomplete, as rows the packet and the workbook both render."""
    gaps: list[dict[str, str]] = []

    def add(kind: str, venue: str, detail: str, action: str) -> None:
        gaps.append({"kind": kind, "venue": venue, "detail": detail, "action": action})

    for r in rows:
        if not r["license_id"]:
            add("License number missing", r["venue_name"],
                f"Screen \"{r['screen_label'] or 'unlabelled'}\" has no license number on file.",
                "Record the license number from the unit on site; we will label it after.")

    seen: set[str] = set()
    for r in rows:
        venue = r["venue_name"]
        if venue in seen:
            continue
        seen.add(venue)

        if not r["address"]:
            add("No address", venue,
                "Not in the host list, the NTV360 dashboard, or the geocode file.",
                "MCTV to supply an address before this stop can be routed.")
        elif r["lat"] is None:
            add("Not geocoded", venue,
                f"Address on file ({r['address']}) has no coordinates, so it is "
                "routed at the end of the day rather than in sequence.",
                "Drive it with the cluster nearest its address.")

        if not r["contact_name"] or r["contact_name"].lower() == "on file":
            add("No host contact", venue,
                "No named contact in the host list.",
                "Ask for the manager on duty and record the name in Notes.")
        elif not r["contact_phone"]:
            add("No contact phone", venue,
                f"{r['contact_name']} is on file with no phone number.",
                "Record a phone number in Notes if the contact gives one.")

        counts = {c for c in (r["sweep_screens"], r["rate_screens"],
                              r["dashboard_licenses"]) if c}
        if len(counts) > 1:
            add("Screen count disagreement", venue,
                f"Whitelist sweep shows {r['sweep_screens'] or 0}, rate model shows "
                f"{r['rate_screens'] or 0}, NTV360 shows {r['dashboard_licenses'] or 0}.",
                "Count the physical MCTV screens at this venue and report the number.")

        if r["same_building"]:
            key = _building_key(r["address"])
            neighbours = sorted({
                x["venue_name"] for x in rows
                if _building_key(x["address"]) == key and x["venue_name"] != venue
            })
            add("Shared address", venue,
                f"Shares a building with {', '.join(neighbours)}.",
                "One trip covers every screen at this address — do them together.")

    return gaps


def roster_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Headline numbers for the page KPIs and the brief's metrics banner."""
    venues = {normalize(r["venue_name"]) for r in rows}
    return {
        "screens": len(rows),
        "venues": len(venues),
        "with_license": sum(1 for r in rows if r["license_id"]),
        "missing_license": sum(1 for r in rows if not r["license_id"]),
        "over_target": sum(1 for r in rows if r["over_target"]),
        "no_address": sum(1 for r in rows if not r["address"]),
        "no_contact": len({
            r["venue_name"] for r in rows
            if not r["contact_name"] or r["contact_name"].lower() == "on file"
        }),
        "stops": len({r["stop_no"] for r in rows if r["stop_no"]}),
    }


# ── Field sheet workbook ─────────────────────────────────────────────────────

def _maps_link(row: dict[str, Any]) -> str:
    if row.get("lat") is not None and row.get("lon") is not None:
        return f"https://www.google.com/maps/search/?api=1&query={row['lat']},{row['lon']}"
    if row.get("address"):
        from urllib.parse import quote
        return f"https://www.google.com/maps/search/?api=1&query={quote(row['address'])}"
    return ""


def build_workbook(rows: list[dict[str, Any]], gaps: list[dict[str, str]],
                   market: str) -> bytes:
    """The field sheet, as .xlsx bytes.

    Four sheets: the capture grid, the route, the known gaps, and a legend.
    Capture columns carry dropdowns so the data comes back clean enough to
    import without hand-cleaning.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    navy = "FF1B1F3B"
    gold = "FFC5A55A"
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
    prefill_font = Font(name="Arial", size=10)
    prefill_fill = PatternFill("solid", fgColor="FFF2F2F5")
    thin = Side(style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Field Sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Field Sheet"

    headers = [h for h, _, _ in PREFILL_COLUMNS] + [h for h, _, _ in CAPTURE_COLUMNS]
    ws.append(headers)
    for col, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=navy if col <= len(PREFILL_COLUMNS) else gold)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32

    for r in rows:
        values = []
        for _, key, _ in PREFILL_COLUMNS:
            value = r.get(key)
            values.append("" if value is None else value)
        ws.append(values + [""] * len(CAPTURE_COLUMNS))

    for idx, (_, _, width) in enumerate(PREFILL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for offset, (_, width, _) in enumerate(CAPTURE_COLUMNS):
        ws.column_dimensions[
            get_column_letter(len(PREFILL_COLUMNS) + 1 + offset)
        ].width = width

    last_row = max(ws.max_row, 2)
    for row_cells in ws.iter_rows(min_row=2, max_row=last_row):
        for col, cell in enumerate(row_cells, start=1):
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col <= len(PREFILL_COLUMNS):
                cell.font = prefill_font
                cell.fill = prefill_fill

    for offset, (header, _, options) in enumerate(CAPTURE_COLUMNS):
        if not options:
            continue
        letter = get_column_letter(len(PREFILL_COLUMNS) + 1 + offset)
        dv = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"', allow_blank=True,
        )
        dv.prompt = CAPTURE_LEGEND.get(header, "")
        dv.promptTitle = header
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    ws.freeze_panes = f"{get_column_letter(len(PREFILL_COLUMNS) + 1)}2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    # ── Route ────────────────────────────────────────────────────────────────
    route = wb.create_sheet("Route")
    route.append(["Stop", "Cluster", "Venue", "Screens here", "Address",
                  "Contact", "Phone", "Shared address?", "Map"])
    for cell in route[1]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    seen_stops: set[int] = set()
    for r in rows:
        if r["stop_no"] in seen_stops:
            continue
        seen_stops.add(r["stop_no"])
        here = sum(1 for x in rows if x["stop_no"] == r["stop_no"])
        route.append([
            r["stop_no"], r["cluster"], r["venue_name"], here, r["address"],
            r["contact_name"], r["contact_phone"],
            "Yes" if r["same_building"] else "", _maps_link(r),
        ])
    for idx, width in enumerate([6, 24, 34, 13, 40, 20, 16, 15, 46], start=1):
        route.column_dimensions[get_column_letter(idx)].width = width
    route.freeze_panes = "A2"

    # ── Gaps ─────────────────────────────────────────────────────────────────
    gap_sheet = wb.create_sheet("Gaps")
    gap_sheet.append(["What is missing", "Venue", "Detail", "What to do"])
    for cell in gap_sheet[1]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for g in gaps:
        gap_sheet.append([g["kind"], g["venue"], g["detail"], g["action"]])
    for idx, width in enumerate([28, 34, 60, 54], start=1):
        gap_sheet.column_dimensions[get_column_letter(idx)].width = width
    for row_cells in gap_sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    gap_sheet.freeze_panes = "A2"

    # ── Legend ───────────────────────────────────────────────────────────────
    legend = wb.create_sheet("Legend")
    legend.append([f"MCTV screen audit — {market_label(market)}"])
    legend["A1"].font = Font(name="Arial", size=13, bold=True)
    legend.append([f"Generated {date.today().isoformat()}"])
    legend.append([])
    legend.append(["Column", "What to record"])
    for cell in legend[4]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=navy)
    for header, _, _ in CAPTURE_COLUMNS:
        legend.append([header, CAPTURE_LEGEND.get(header, "")])
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 96
    for row_cells in legend.iter_rows(min_row=5):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Reading a completed field sheet back in ──────────────────────────────────

def parse_completed_workbook(source, market: str | None = None) -> list[dict[str, Any]]:
    """Read a returned field sheet into ``screen_assets``-shaped rows.

    ``source`` is a path or any file-like object (a Streamlit upload works).
    Rows where nothing was captured are skipped — a technician who did not
    reach a screen should not overwrite what we already know about it.

    Columns are matched by header text, so a contractor inserting a column or
    reordering does not break the import.
    """
    from openpyxl import load_workbook

    wb = load_workbook(source, data_only=True)
    ws = wb["Field Sheet"] if "Field Sheet" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    index = {name: i for i, name in enumerate(headers) if name}

    def value_at(row, header):
        pos = index.get(header)
        if pos is None or pos >= len(row):
            return None
        val = row[pos]
        if isinstance(val, str):
            val = val.strip()
        return val or None

    out: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        captured = {
            column: value_at(raw, header)
            for header, column in CAPTURE_DB_FIELDS.items()
            if value_at(raw, header) is not None
        }
        if not captured:
            continue

        record: dict[str, Any] = {
            "license_id": value_at(raw, "License number"),
            "venue_name": value_at(raw, "Venue") or "",
            "short_code": value_at(raw, "Short code"),
            "screen_label": value_at(raw, "Screen"),
            "market": market,
        }
        record.update(captured)

        # Excel hands back dates as datetimes and sizes as text like '43"'.
        if record.get("audited_at") is not None:
            record["audited_at"] = _as_iso_date(record["audited_at"])
        if record.get("display_size_in") is not None:
            record["display_size_in"] = _as_number(record["display_size_in"])

        record = {k: v for k, v in record.items() if v is not None}
        if record.get("venue_name"):
            out.append(record)
    return out


def _as_iso_date(value) -> str | None:
    from datetime import date as _date, datetime as _datetime
    if isinstance(value, _datetime):
        return value.date().isoformat()
    if isinstance(value, _date):
        return value.isoformat()
    text = str(value).strip()[:10]
    try:
        return _datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None


def _as_number(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    digits = "".join(ch for ch in str(value) if ch.isdigit() or ch == ".")
    try:
        return float(digits) if digits else None
    except ValueError:
        return None


def save_assets(records: list[dict[str, Any]]) -> tuple[int, list[str]]:
    """Write parsed field-sheet rows to ``screen_assets``.

    Upserts on ``license_id`` so re-importing a corrected sheet updates the
    same row rather than duplicating it. Rows with no license number yet are
    inserted, since there is nothing to key on until the number is known.

    Returns ``(saved_count, errors)``.
    """
    from services.supabase_client import insert_row, upsert_row

    saved, errors = 0, []
    for record in records:
        label = record.get("short_code") or record.get("venue_name") or "unknown screen"
        try:
            if record.get("license_id"):
                ok = upsert_row("screen_assets", record, on_conflict="license_id")
            else:
                ok = insert_row("screen_assets", record)
            if ok:
                saved += 1
            else:
                errors.append(f"{label}: write rejected by the database")
        except Exception as exc:  # pragma: no cover - network/permission issues
            errors.append(f"{label}: {exc}")
    return saved, errors
