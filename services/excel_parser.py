"""NTV360 Excel export parser with auto-detection for 3 formats."""

import re
import openpyxl
from pathlib import Path
from models.report_data import PlayRecord, VenueRecord, TractionReportInput, CategoryBreakdown


# ── Venue categorization rules ──────────────────────────────────────────────
# Each rule is (regex_pattern, category_name). First match wins.
VENUE_RULES = [
    (r"restaurant|grill|bbq|barbeque|barbecue|cantina|seafood|bar\b|kitchen|"
     r"caf[eé]|coffee|nutrition|chicken|pizza|taco|wing|burger|diner|bistro|"
     r"bakery|brewing|steakhouse|sushi|pho|ramen|thai|sub[s ]|sandwich|waffle|"
     r"donut|deli\b|ice\s*cream|smoothie|juice|pub\b|tavern|eatery|catfish|"
     r"crawfish|stick\b|\bcoop\b|mexicana|mexican|casa\b|trattoria|hibachi|"
     r"buffet|wok\b|noodle|dumpling|poke\b|bowl\b.*kitchen",
     "Restaurant & Bar"),
    (r"salon|beauty|aesthetics|waxing|hair|styling|tan\b|barber|nails?\b|lash|"
     r"spa\b|brow|fade|cutz|cuts\b|cosmetic|beard|style\s+society",
     "Salon & Beauty"),
    (r"medical|dental|clinic|hospital|cardiology|urgent|doctor|optom|chiro|"
     r"physical.?therap|derma|ortho|pharmacy|health\b|pediatr|family\s+med|"
     r"wellness|vision|eye\b|allergy|asthma|medicine|internal\s+med|"
     r"testosterone|hormone|iv\s+therap|infusion",
     "Medical & Dental"),
    (r"chevrolet|collision|auto\b|tire|lube|car\s*wash|motor\b|toyota|\bford\b|"
     r"\bhonda\b|\bnissan\b|hyundai|\bkia\b|\bjeep\b|dodge|chrysler|chevy|bmw|mercedes|"
     r"mechanic|body\s+shop|muffler|transmission",
     "Auto & Service"),
    (r"gym|fitness|crossfit|yoga|martial|boxing|athletic|weight|muscle|"
     r"training|workout|body\b.*fit|pilates|spin\b|performance\b.*(?:fitness|gym)",
     "Health & Fitness"),
    (r"liquor|wine|spirit|beer|package\s+store|beverage",
     "Liquor & Spirits"),
    (r"school|university|college|education|tutor|academy|learning|daycare|"
     r"child\s*care|montessori",
     "Education"),
    (r"insurance|law\b|legal|accounting|real\s+estate|office|bank|financ|"
     r"consult|mortgage|tax\b|cpa\b|title\b|escrow|invest|technologies|"
     r"tech\b|solutions\b|systems\b|exceed",
     "Professional Services"),
    (r"hotel|boutique|shop\b|retail|florist|flower|jewel|gift|thrift|cloth|"
     r"fashion|dress|bridal|pet\b|feed\b|hardware|supply",
     "Retail & Hospitality"),
    (r"skate|commission|chamber|church|non.?profit|community|library|museum|"
     r"civic|recreation|park\b|ymca|rotary|kiwanis|elks|vfw|legion",
     "Community & Entertainment"),
]


def classify_venue(name: str) -> str:
    """Classify a venue into a business category based on its name.

    Uses keyword regex matching against VENUE_RULES. First match wins.
    Returns 'General' if no rule matches.
    """
    lower = name.lower()
    for pattern, category in VENUE_RULES:
        if re.search(pattern, lower):
            return category
    return "General"


def format_date_range(start: str, end: str) -> str:
    """Format a date range as 'January 31 – February 23, 2026'.

    Handles common date formats from NTV360 exports:
      - 2026-01-31, 01/31/2026, 2026/01/31

    Returns the raw 'start - end' string if parsing fails.
    """
    from datetime import datetime as dt

    def _parse_date(s: str):
        s = str(s).strip()[:10]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%m-%d-%Y"):
            try:
                return dt.strptime(s, fmt)
            except ValueError:
                continue
        return None

    d1 = _parse_date(start)
    d2 = _parse_date(end)

    if not d1 or not d2:
        # Fallback: return raw string
        return f"{start} - {end}" if start and end else start or end or ""

    # Same year: "January 31 – February 23, 2026"
    if d1.year == d2.year:
        return f"{d1.strftime('%B %d').lstrip('0')} \u2013 {d2.strftime('%B %d, %Y').lstrip('0')}"
    else:
        return f"{d1.strftime('%B %d, %Y')} \u2013 {d2.strftime('%B %d, %Y')}"


def parse_duration(text) -> int:
    """Parse NTV360 duration values to total seconds.

    Handles:
      - "8h 17m 40s" format (NTV360 text export)
      - "8:17:40" / "0:17:40" (HH:MM:SS from Excel time cells)
      - datetime.timedelta objects (openpyxl returns these for duration cells)
      - datetime.time objects (openpyxl returns these for time-formatted cells)
    """
    from datetime import timedelta, time as dt_time

    if not text:
        return 0

    # Handle timedelta objects directly (openpyxl duration cells)
    if isinstance(text, timedelta):
        return int(text.total_seconds())

    # Handle datetime.time objects (openpyxl time-formatted cells)
    if isinstance(text, dt_time):
        return text.hour * 3600 + text.minute * 60 + text.second

    text = str(text).strip()
    if not text:
        return 0

    # Try "Xh Ym Zs" format
    h_match = re.search(r"(\d+)\s*h", text)
    m_match = re.search(r"(\d+)\s*m", text)
    s_match = re.search(r"(\d+)\s*s", text)
    if h_match or m_match or s_match:
        hours = int(h_match.group(1)) if h_match else 0
        minutes = int(m_match.group(1)) if m_match else 0
        seconds = int(s_match.group(1)) if s_match else 0
        return hours * 3600 + minutes * 60 + seconds

    # Try "HH:MM:SS" or "H:MM:SS" format (Excel time cells stringified)
    colon_match = re.match(r"(\d+):(\d{1,2}):(\d{1,2})", text)
    if colon_match:
        return (int(colon_match.group(1)) * 3600 +
                int(colon_match.group(2)) * 60 +
                int(colon_match.group(3)))

    # Try "MM:SS" format
    short_match = re.match(r"(\d+):(\d{1,2})$", text)
    if short_match:
        return int(short_match.group(1)) * 60 + int(short_match.group(2))

    return 0


def format_duration(seconds: int) -> str:
    """Format seconds back to 'Xh Ym' string."""
    if seconds <= 0:
        return "0h 0m"
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    if hours > 0:
        return f"{hours:,}h {minutes}m"
    return f"{minutes}m"


def parse_network_dashboard(file_path) -> dict:
    """Parse the MCTV Network Dashboard Excel file.

    Returns a dict mapping host_name (lowercase) -> {
        host_name, category, general_category, address,
        license_count, traffic, dwell_time, impressions
    }

    The dashboard has per-venue data including monthly foot traffic estimates,
    average dwell time (minutes), and estimated monthly impressions.
    Formula: Impressions = Traffic × Dwell Time × License Count / 15

    file_path can be a string path or a file-like object (from Streamlit uploader).
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)

    # Use the "All MCTV Hosts" sheet if available, else first sheet
    if "All MCTV Hosts" in wb.sheetnames:
        ws = wb["All MCTV Hosts"]
    else:
        ws = wb.active

    # Find header row and build column map
    col_map = {}
    header_row = 1
    for c in range(1, min((ws.max_column or 10) + 1, 30)):
        hdr = str(ws.cell(1, c).value or "").strip().lower()
        if not hdr:
            continue
        if hdr in ("host name", "host"):
            col_map["host"] = c
        elif hdr == "category":
            col_map["category"] = c
        elif "general" in hdr and "category" in hdr:
            col_map["general_category"] = c
        elif hdr == "address":
            col_map["address"] = c
        elif "license" in hdr or "screen" in hdr:
            col_map["licenses"] = c
        elif "traffic" in hdr:
            col_map["traffic"] = c
        elif "dwell" in hdr:
            col_map["dwell"] = c
        elif "impression" in hdr:
            col_map["impressions"] = c
        elif hdr in ("dealer name", "dealer"):
            col_map["dealer"] = c

    if "host" not in col_map:
        return {}

    lookup = {}
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        host = str(ws.cell(r, col_map["host"]).value or "").strip()
        if not host:
            continue
        # Skip summary rows
        if host.lower() in ("sum", "ave", "total", "average"):
            continue

        def _float(col_key):
            if col_key not in col_map:
                return 0.0
            val = ws.cell(r, col_map[col_key]).value
            try:
                return float(val) if val else 0.0
            except (ValueError, TypeError):
                return 0.0

        def _int(col_key):
            if col_key not in col_map:
                return 1
            val = ws.cell(r, col_map[col_key]).value
            try:
                return int(val) if val else 1
            except (ValueError, TypeError):
                return 1

        def _str(col_key):
            if col_key not in col_map:
                return ""
            return str(ws.cell(r, col_map[col_key]).value or "").strip()

        entry = {
            "host_name": host,
            "category": _str("category"),
            "general_category": _str("general_category"),
            "address": _str("address"),
            "license_count": _int("licenses"),
            "traffic": _float("traffic"),
            "dwell_time": _float("dwell"),
            "impressions": _float("impressions"),
        }

        lookup[host.lower()] = entry

    return lookup


def enrich_report_with_dashboard(data, dashboard_lookup: dict):
    """Enrich a TractionReportInput with data from the Network Dashboard.

    Matches venue records by host name and populates:
      - monthly_traffic, dwell_time_minutes, monthly_impressions, screen_count, address
    Also computes totals on the report data object.
    """
    if not dashboard_lookup:
        return

    total_impressions = 0.0
    total_traffic = 0.0
    dwell_sum = 0.0
    dwell_count = 0

    for venue in data.venue_records:
        key = venue.host_name.lower()
        entry = dashboard_lookup.get(key)
        if entry:
            venue.monthly_traffic = entry["traffic"]
            venue.dwell_time_minutes = entry["dwell_time"]
            venue.monthly_impressions = entry["impressions"]
            venue.screen_count = entry["license_count"]
            if not venue.address:
                venue.address = entry["address"]
            total_impressions += entry["impressions"]
            total_traffic += entry["traffic"]
            if entry["dwell_time"] > 0:
                dwell_sum += entry["dwell_time"]
                dwell_count += 1

    data.total_impressions = total_impressions
    data.total_monthly_traffic = total_traffic
    if dwell_count > 0:
        data.avg_dwell_time = round(dwell_sum / dwell_count, 1)


def detect_format(wb: openpyxl.Workbook) -> str:
    """Auto-detect which NTV360 export format this workbook uses.

    Returns: 'per_content', 'content_report', 'traction_report', or 'unknown'
    """
    sheet = wb.active

    # Check first cell for clues
    a1 = str(sheet.cell(1, 1).value or "").strip()
    a2 = str(sheet.cell(2, 1).value or "").strip()
    a5 = str(sheet.cell(5, 1).value or "").strip().lower()

    # Format 2: Per-content report -- Row 1 starts with "Filename"
    if a1.lower().startswith("filename"):
        return "per_content"

    # Format 2 alt: Row 1 has a .webm filename
    if ".webm" in a1.lower() or "476mctv" in a1.lower():
        return "per_content"

    # Format 3: Pre-formatted traction report -- has "PLAY COUNT" or "TRACTION"
    if "play count" in a1.lower() or "traction" in a1.lower():
        return "traction_report"

    # Check if it's a prepared traction report by looking for typical headers
    for row in range(1, 10):
        for col in range(1, 10):
            val = str(sheet.cell(row, col).value or "").strip().lower()
            if val in ("host location", "host name", "host"):
                next_val = str(sheet.cell(row, col + 1).value or "").strip().lower()
                if "play" in next_val or "air" in next_val or "total" in next_val:
                    return "traction_report"

    # Format 1: Content report -- has date range sheet name or Content Name header
    for name in wb.sheetnames:
        if re.match(r"\d{4}-\d{2}-\d{2}", name):
            return "content_report"

    # Check for content report headers
    if a5 == "host" or "content" in a1.lower():
        return "content_report"

    # Check for NTV360 host list
    if "host name" in a1.lower() or "dealer" in a1.lower():
        return "host_list"

    return "unknown"


def parse_per_content_report(wb: openpyxl.Workbook) -> list:
    """Parse per-content report format (NTV360 export).

    NTV360 exports use this structure:
        Row 1: Filename | [content_filename]
        Row 2: [blank] | [blank] | Total Count | Total Duration
        Row 3: [blank] | [blank] | [total_plays]  | [total_duration_str]
        Row 4: [empty]
        Row 5: Host | City | State | Zip Code | Region | Playlist | Play Count | Play Duration | Start Date | End Date
        Row 6+: venue data rows

    Column positions vary across exports (6-col or 10-col), so we use
    header-name-based mapping instead of hardcoded indices.
    """
    records = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        if max_row is None or max_row < 5:
            continue

        # Get content name from row 1
        content_name = ""
        a1 = str(sheet.cell(1, 1).value or "")
        b1 = str(sheet.cell(1, 2).value or "")
        if "filename" in a1.lower():
            content_name = b1.strip()
        else:
            content_name = a1.strip()

        # Find the header row (usually row 5) — look for "host" in column A
        header_row = None
        for r in range(3, min(15, max_row + 1)):
            val = str(sheet.cell(r, 1).value or "").strip().lower()
            if val in ("host", "host name", "host location"):
                header_row = r
                break

        if header_row is None:
            continue

        # Build column map from header names (flexible for 6-col or 10-col)
        col_map = {"host": 1}  # We already know host is col 1
        max_col = sheet.max_column or 10
        for c in range(2, max_col + 1):
            hdr = str(sheet.cell(header_row, c).value or "").strip().lower()
            if not hdr:
                continue
            if hdr == "city":
                col_map["city"] = c
            elif hdr == "playlist":
                col_map["playlist"] = c
            elif "play count" in hdr or hdr == "plays":
                col_map["plays"] = c
            elif "play duration" in hdr or "duration" in hdr:
                col_map["duration"] = c
            elif "start" in hdr:
                col_map["start"] = c
            elif "end" in hdr:
                col_map["end"] = c

        # Fallback: if no "plays" column found, try positional (old 6-col)
        if "plays" not in col_map:
            num_cols = sum(1 for c in range(1, 11) if sheet.cell(header_row, c).value)
            if num_cols >= 6:
                col_map.setdefault("plays", 3)
                col_map.setdefault("duration", 4)
                col_map.setdefault("start", 5)
                col_map.setdefault("end", 6)
            else:
                col_map.setdefault("plays", 2)
                col_map.setdefault("duration", 3)

        # Parse data rows
        for r in range(header_row + 1, max_row + 1):
            host = str(sheet.cell(r, col_map["host"]).value or "").strip()
            if not host or host.lower().startswith("total") or host.lower().startswith("sum"):
                continue

            # Read play count
            raw_plays = sheet.cell(r, col_map.get("plays", 2)).value
            try:
                play_count = int(raw_plays) if raw_plays is not None else 0
            except (ValueError, TypeError):
                play_count = 0

            # Read duration — keep raw value for parse_duration (handles timedelta)
            raw_duration = sheet.cell(r, col_map.get("duration", 3)).value
            duration_seconds = parse_duration(raw_duration)
            duration_str = format_duration(duration_seconds) if duration_seconds else str(raw_duration or "")

            # Read optional fields
            city = str(sheet.cell(r, col_map["city"]).value or "").strip() if "city" in col_map else ""
            playlist = str(sheet.cell(r, col_map["playlist"]).value or "").strip() if "playlist" in col_map else ""
            start = str(sheet.cell(r, col_map["start"]).value or "") if "start" in col_map else ""
            end = str(sheet.cell(r, col_map["end"]).value or "") if "end" in col_map else ""

            # Skip demo/test venues (Spec B-1.3)
            if "demo" in host.lower() or "demo" in playlist.lower():
                continue

            records.append(PlayRecord(
                host_name=host,
                content_name=content_name,
                play_count=play_count,
                play_duration_str=duration_str,
                play_duration_seconds=duration_seconds,
                playlist=playlist,
                city=city,
                start_date=start,
                end_date=end,
            ))

    return records


def parse_content_report(wb: openpyxl.Workbook) -> list:
    """Parse the content report format with multiple content pieces per host."""
    records = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        if max_row is None or max_row < 2:
            continue

        # Find header row
        header_row = None
        for r in range(1, min(10, max_row + 1)):
            val = str(sheet.cell(r, 1).value or "").strip().lower()
            if "content" in val or "host" in val:
                header_row = r
                break

        if header_row is None:
            header_row = 1

        for r in range(header_row + 1, max_row + 1):
            content_name = str(sheet.cell(r, 1).value or "").strip()
            host_name = str(sheet.cell(r, 2).value or "").strip()
            play_count = sheet.cell(r, 3).value or 0
            raw_duration = sheet.cell(r, 4).value

            if not host_name:
                continue

            # Skip demo/test venues
            if "demo" in host_name.lower():
                continue

            try:
                play_count = int(play_count)
            except (ValueError, TypeError):
                play_count = 0

            duration_seconds = parse_duration(raw_duration)
            duration_str = format_duration(duration_seconds) if duration_seconds else str(raw_duration or "")

            records.append(PlayRecord(
                host_name=host_name,
                content_name=content_name,
                play_count=play_count,
                play_duration_str=duration_str,
                play_duration_seconds=duration_seconds,
            ))

    return records


def parse_traction_report(wb: openpyxl.Workbook) -> list:
    """Parse pre-formatted traction report Excel files."""
    records = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        if max_row is None or max_row < 3:
            continue

        # Find header row by looking for "Host" or "Host Name" or "Host Location"
        header_row = None
        host_col = None
        for r in range(1, min(15, max_row + 1)):
            for c in range(1, min(10, (sheet.max_column or 5) + 1)):
                val = str(sheet.cell(r, c).value or "").strip().lower()
                if val in ("host", "host name", "host location"):
                    header_row = r
                    host_col = c
                    break
            if header_row:
                break

        if header_row is None or host_col is None:
            continue

        # Map column headers
        col_map = {}
        for c in range(1, (sheet.max_column or 10) + 1):
            val = str(sheet.cell(header_row, c).value or "").strip().lower()
            if "play count" in val or "ad plays" in val or "plays" in val:
                col_map["plays"] = c
            elif "duration" in val or "air time" in val:
                col_map["duration"] = c
            elif "screen" in val or "license" in val:
                col_map["screens"] = c
            elif "category" in val:
                col_map["category"] = c
            elif val == "city":
                col_map["city"] = c
            elif "address" in val:
                col_map["address"] = c
            elif "traffic" in val:
                col_map["traffic"] = c
            elif "dwell" in val:
                col_map["dwell"] = c
            elif "impression" in val:
                col_map["impressions"] = c
            elif "playlist" in val:
                col_map["playlist"] = c

        # Parse data rows
        for r in range(header_row + 1, max_row + 1):
            host = str(sheet.cell(r, host_col).value or "").strip()
            if not host or host.lower().startswith("total") or host.lower().startswith("sum"):
                continue

            # Skip demo/test venues
            playlist = str(sheet.cell(r, col_map["playlist"]).value or "").strip() if "playlist" in col_map else ""
            if "demo" in host.lower() or "demo" in playlist.lower():
                continue

            play_count = 0
            if "plays" in col_map:
                raw = sheet.cell(r, col_map["plays"]).value
                try:
                    play_count = int(raw) if raw else 0
                except (ValueError, TypeError):
                    play_count = 0

            raw_duration = None
            if "duration" in col_map:
                raw_duration = sheet.cell(r, col_map["duration"]).value
            duration_seconds = parse_duration(raw_duration)
            duration_str = format_duration(duration_seconds) if duration_seconds else str(raw_duration or "")

            city = str(sheet.cell(r, col_map["city"]).value or "").strip() if "city" in col_map else ""

            records.append(PlayRecord(
                host_name=host,
                play_count=play_count,
                play_duration_str=duration_str,
                play_duration_seconds=duration_seconds,
                playlist=playlist,
                city=city,
            ))

    return records


def parse_excel(file_path) -> list:
    """Auto-detect format and parse an NTV360 Excel file.

    file_path can be a string path or a file-like object (from Streamlit uploader).
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)

    fmt = detect_format(wb)

    if fmt == "per_content":
        return parse_per_content_report(wb)
    elif fmt == "content_report":
        return parse_content_report(wb)
    elif fmt == "traction_report":
        return parse_traction_report(wb)
    else:
        # Try traction report as fallback
        records = parse_traction_report(wb)
        if records:
            return records
        return parse_per_content_report(wb)


def aggregate_by_host(records: list) -> dict:
    """Aggregate play records by host name.

    Returns dict mapping host_name -> {total_plays, total_seconds, content_names, city, ...}
    """
    hosts = {}
    for rec in records:
        name = rec.host_name
        if name not in hosts:
            hosts[name] = {
                "host_name": name,
                "total_plays": 0,
                "total_seconds": 0,
                "content_names": set(),
                "playlists": set(),
                "city": rec.city or "",
                "first_date": rec.start_date,
                "last_date": rec.end_date,
            }
        hosts[name]["total_plays"] += rec.play_count
        hosts[name]["total_seconds"] += rec.play_duration_seconds
        if rec.content_name:
            hosts[name]["content_names"].add(rec.content_name)
        if rec.playlist:
            hosts[name]["playlists"].add(rec.playlist)
        # Keep the first non-empty city we see
        if rec.city and not hosts[name]["city"]:
            hosts[name]["city"] = rec.city
        if rec.start_date and (not hosts[name]["first_date"] or rec.start_date < hosts[name]["first_date"]):
            hosts[name]["first_date"] = rec.start_date
        if rec.end_date and (not hosts[name]["last_date"] or rec.end_date > hosts[name]["last_date"]):
            hosts[name]["last_date"] = rec.end_date

    # Convert sets to lists for serialization
    for host_data in hosts.values():
        host_data["content_names"] = list(host_data["content_names"])
        host_data["playlists"] = list(host_data["playlists"])
        host_data["total_air_time"] = format_duration(host_data["total_seconds"])

    return hosts


def build_report_data(records: list, advertiser_name: str,
                      campaign_period: str = "") -> TractionReportInput:
    """Build a TractionReportInput from raw play records."""
    hosts = aggregate_by_host(records)

    # Exclude demo/test venues at aggregate level as a safety net
    # (parsers also filter, but this catches any that slip through)
    hosts = {
        name: data for name, data in hosts.items()
        if "demo" not in name.lower()
        and not any("demo" in p.lower() for p in data.get("playlists", []))
    }

    total_plays = sum(h["total_plays"] for h in hosts.values())
    total_seconds = sum(h["total_seconds"] for h in hosts.values())

    # Build venue records sorted by total plays descending
    venue_records = []
    for host_data in sorted(hosts.values(), key=lambda x: x["total_plays"], reverse=True):
        pct = (host_data["total_plays"] / total_plays * 100) if total_plays > 0 else 0
        venue_records.append(VenueRecord(
            host_name=host_data["host_name"],
            city=host_data.get("city", ""),
            business_category=classify_venue(host_data["host_name"]),
            total_plays=host_data["total_plays"],
            total_air_time=host_data["total_air_time"],
            first_aired=host_data.get("first_date", ""),
            last_aired=host_data.get("last_date", ""),
            pct_of_total=round(pct, 1),
        ))

    # Auto-detect campaign period from date range in data
    auto_start = ""
    auto_end = ""
    for h in hosts.values():
        s = h.get("first_date", "")
        e = h.get("last_date", "")
        if s and (not auto_start or s < auto_start):
            auto_start = s
        if e and (not auto_end or e > auto_end):
            auto_end = e

    return TractionReportInput(
        advertiser_name=advertiser_name,
        campaign_period=campaign_period,
        campaign_start=auto_start,
        campaign_end=auto_end,
        venue_records=venue_records,
        total_plays=total_plays,
        total_screen_count=len(hosts),
        total_air_time=format_duration(total_seconds),
    )
